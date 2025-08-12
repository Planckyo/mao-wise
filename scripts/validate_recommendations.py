#!/usr/bin/env python3
"""
推荐验证脚本 - 文献对照验证

将批量推荐方案与KB文献进行逐条匹配，输出历史先例分析、最相近文献和参数差异摘要。

功能特性：
- 批量方案与知识库文献的相似度匹配
- 基于体系+关键电参数的智能检索
- 生成详细的验证报告（Excel格式）
- 提供命中率和相似度统计分析
- 支持自定义相似度阈值和TopK设置

使用示例：
python scripts/validate_recommendations.py --plans tasks/batch_20250812_2246/plans.csv --kb datasets/index_store --topk 3
python scripts/validate_recommendations.py --plans tasks/batch_*/plans.csv --kb datasets/index_store --topk 5 --threshold 0.7
"""

import argparse
import csv
import json
import sys
import pathlib
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
import logging
from datetime import datetime

# 确保能找到maowise包
REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from maowise.utils.logger import logger
from maowise.kb.search import kb_search
from maowise.utils.config import load_config

@dataclass
class ValidationResult:
    """单个方案的验证结果"""
    plan_id: str
    system: str
    match_found: bool
    similarity_score: float
    nearest_citations: List[Dict[str, Any]]
    delta_params: Dict[str, float]
    query_used: str
    validation_time: str

@dataclass
class ValidationSummary:
    """验证摘要统计"""
    total_plans: int
    matched_plans: int
    unmatched_plans: int
    match_rate: float
    avg_similarity: float
    avg_delta_voltage: float
    avg_delta_current: float
    most_cited_sources: List[Tuple[str, int]]
    validation_time: str

class RecommendationValidator:
    """推荐验证器"""
    
    def __init__(self, kb_path: str, similarity_threshold: float = 0.6):
        self.kb_path = pathlib.Path(kb_path)
        self.similarity_threshold = similarity_threshold
        self.config = load_config()
        
    def _construct_query(self, plan: Dict[str, Any]) -> str:
        """根据方案构造检索查询"""
        system = plan.get('system', '').lower()
        
        # 从方案中提取参数，如果没有则从YAML文件中获取
        plan_params = self._extract_plan_params(plan)
        voltage = plan_params.get('voltage_V', 0)
        current = plan_params.get('current_density_A_dm2', 0)
        frequency = plan_params.get('frequency_Hz', 0)
        duty_cycle = plan_params.get('duty_cycle_pct', 0)
        time_min = plan_params.get('time_min', 0)
        
        # 构造查询：体系 + 关键电参数
        query_parts = []
        
        # 体系信息
        if system == 'silicate':
            query_parts.append("silicate Na2SiO3 alkaline electrolyte")
        elif system == 'zirconate':
            query_parts.append("zirconate K2ZrF6 fluoride electrolyte")
        else:
            query_parts.append("micro arc oxidation electrolyte")
        
        # 电参数（选择最重要的2个）
        if voltage > 0:
            query_parts.append(f"{voltage:.0f}V voltage")
        if current > 0:
            query_parts.append(f"{current:.1f}A/dm2 current density")
        
        # 可选参数
        if frequency > 0:
            query_parts.append(f"{frequency:.0f}Hz frequency")
        if duty_cycle > 0:
            query_parts.append(f"{duty_cycle:.0f}% duty cycle")
        if time_min > 0:
            query_parts.append(f"{time_min:.0f}min time")
        
        # 限制查询长度，选择top-2关键参数
        if len(query_parts) > 3:
            query_parts = query_parts[:3]  # 体系 + 2个关键参数
            
        return " ".join(query_parts)
    
    def _extract_plan_params(self, plan: Dict[str, Any]) -> Dict[str, float]:
        """提取方案的关键参数"""
        # 如果CSV中没有参数字段，尝试从YAML文件中解析
        params = {
            'voltage_V': float(plan.get('voltage_V', plan.get('voltage', 0))),
            'current_density_A_dm2': float(plan.get('current_density_A_dm2', plan.get('current_density', 0))),
            'frequency_Hz': float(plan.get('frequency_Hz', plan.get('frequency', 0))),
            'duty_cycle_pct': float(plan.get('duty_cycle_pct', plan.get('duty_cycle', 0))),
            'time_min': float(plan.get('time_min', plan.get('time', 0)))
        }
        
        # 如果参数都是0，尝试从plan_id推断YAML文件路径并解析
        if all(v == 0 for v in params.values()):
            plan_id = plan.get('plan_id', '')
            if plan_id:
                # 推断YAML文件路径
                batch_dir = pathlib.Path(f"tasks/{plan_id.split('_plan_')[0]}")
                yaml_file = batch_dir / "plans_yaml" / f"{plan_id}.yaml"
                
                if yaml_file.exists():
                    try:
                        import yaml
                        with open(yaml_file, 'r', encoding='utf-8') as f:
                            yaml_content = yaml.safe_load(f) or {}
                        
                        # 从YAML描述中解析参数（简化版）
                        description = yaml_content.get('description', '')
                        if description:
                            extracted = self._extract_citation_params(description)
                            for key in params:
                                if extracted.get(key, 0) > 0:
                                    params[key] = extracted[key]
                    except Exception as e:
                        logger.warning(f"Failed to parse YAML for {plan_id}: {e}")
        
        return params
    
    def _extract_citation_params(self, citation_text: str) -> Dict[str, float]:
        """从文献片段中提取参数（简化版规则抽取）"""
        import re
        
        params = {
            'voltage_V': 0.0,
            'current_density_A_dm2': 0.0,
            'frequency_Hz': 0.0,
            'duty_cycle_pct': 0.0,
            'time_min': 0.0
        }
        
        # 电压匹配
        voltage_match = re.search(r'(\d+(?:\.\d+)?)\s*V', citation_text, re.IGNORECASE)
        if voltage_match:
            params['voltage_V'] = float(voltage_match.group(1))
        
        # 电流密度匹配
        current_match = re.search(r'(\d+(?:\.\d+)?)\s*A/dm[²2]', citation_text, re.IGNORECASE)
        if current_match:
            params['current_density_A_dm2'] = float(current_match.group(1))
        
        # 频率匹配
        freq_match = re.search(r'(\d+(?:\.\d+)?)\s*Hz', citation_text, re.IGNORECASE)
        if freq_match:
            params['frequency_Hz'] = float(freq_match.group(1))
        
        # 占空比匹配
        duty_match = re.search(r'(\d+(?:\.\d+)?)\s*%\s*duty', citation_text, re.IGNORECASE)
        if duty_match:
            params['duty_cycle_pct'] = float(duty_match.group(1))
        
        # 时间匹配
        time_match = re.search(r'(\d+(?:\.\d+)?)\s*min', citation_text, re.IGNORECASE)
        if time_match:
            params['time_min'] = float(time_match.group(1))
        
        return params
    
    def _calculate_param_delta(self, plan_params: Dict[str, float], citation_params: Dict[str, float]) -> Dict[str, float]:
        """计算参数差异百分比"""
        deltas = {}
        
        for param, plan_val in plan_params.items():
            citation_val = citation_params.get(param, 0)
            
            if plan_val > 0 and citation_val > 0:
                delta_pct = ((plan_val - citation_val) / citation_val) * 100
                deltas[param] = round(delta_pct, 1)
            elif plan_val > 0 or citation_val > 0:
                deltas[param] = float('inf')  # 一个有值一个没有
            else:
                deltas[param] = 0.0  # 都没有值
        
        return deltas
    
    def _determine_match(self, similarity_score: float, plan: Dict[str, Any], citations: List[Dict[str, Any]]) -> bool:
        """判断是否找到匹配的历史先例"""
        # 方法1: 相似度阈值
        if similarity_score >= self.similarity_threshold:
            return True
        
        # 方法2: 体系+关键电参数匹配
        plan_system = plan.get('system', '').lower()
        
        for citation in citations:
            citation_text = citation.get('text', '').lower()
            
            # 检查体系匹配
            system_match = False
            if plan_system == 'silicate' and ('silicate' in citation_text or 'na2sio3' in citation_text):
                system_match = True
            elif plan_system == 'zirconate' and ('zirconate' in citation_text or 'k2zrf6' in citation_text):
                system_match = True
            elif 'micro arc' in citation_text or 'mao' in citation_text:
                system_match = True
            
            if system_match:
                # 检查是否包含至少2个关键电参数
                param_count = 0
                if 'v' in citation_text and any(char.isdigit() for char in citation_text):
                    param_count += 1
                if 'a/dm' in citation_text:
                    param_count += 1
                if 'hz' in citation_text:
                    param_count += 1
                
                if param_count >= 2:
                    return True
        
        return False
    
    def validate_plan(self, plan: Dict[str, Any], topk: int = 3) -> ValidationResult:
        """验证单个方案"""
        plan_id = plan.get('plan_id', 'unknown')
        system = plan.get('system', 'unknown')
        
        # 构造查询
        query = self._construct_query(plan)
        
        try:
            # 执行检索
            search_results = kb_search(query, k=topk)
            
            # 处理检索结果
            citations = []
            max_similarity = 0.0
            
            if search_results:
                # kb_search直接返回结果列表
                results_list = search_results if isinstance(search_results, list) else search_results.get('results', [])
                
                for result in results_list[:topk]:
                    similarity = result.get('score', 0.0)
                    max_similarity = max(max_similarity, similarity)
                    
                    citation = {
                        'source': result.get('source', result.get('doc_id', 'unknown')),
                        'page': result.get('page', result.get('chunk_id', 0)),
                        'score': round(similarity, 3),
                        'text': result.get('text', result.get('content', ''))[:200] + '...' if len(result.get('text', result.get('content', ''))) > 200 else result.get('text', result.get('content', ''))
                    }
                    citations.append(citation)
            
            # 判断是否匹配
            match_found = self._determine_match(max_similarity, plan, citations)
            
            # 计算参数差异（与最相近文献比较）
            delta_params = {}
            if citations:
                plan_params = self._extract_plan_params(plan)
                citation_params = self._extract_citation_params(citations[0]['text'])
                delta_params = self._calculate_param_delta(plan_params, citation_params)
            
            return ValidationResult(
                plan_id=plan_id,
                system=system,
                match_found=match_found,
                similarity_score=max_similarity,
                nearest_citations=citations,
                delta_params=delta_params,
                query_used=query,
                validation_time=datetime.now().isoformat()
            )
            
        except Exception as e:
            logger.error(f"Validation failed for plan {plan_id}: {e}")
            return ValidationResult(
                plan_id=plan_id,
                system=system,
                match_found=False,
                similarity_score=0.0,
                nearest_citations=[],
                delta_params={},
                query_used=query,
                validation_time=datetime.now().isoformat()
            )
    
    def validate_batch(self, plans_csv: str, topk: int = 3) -> Tuple[List[ValidationResult], ValidationSummary]:
        """验证整个批次"""
        plans_path = pathlib.Path(plans_csv)
        
        if not plans_path.exists():
            raise FileNotFoundError(f"Plans file not found: {plans_csv}")
        
        # 读取方案CSV
        plans = []
        with open(plans_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            plans = list(reader)
        
        logger.info(f"Validating {len(plans)} plans from {plans_csv}")
        
        # 验证每个方案
        results = []
        for i, plan in enumerate(plans, 1):
            logger.info(f"Validating plan {i}/{len(plans)}: {plan.get('plan_id', 'unknown')}")
            result = self.validate_plan(plan, topk=topk)
            results.append(result)
        
        # 生成摘要统计
        summary = self._generate_summary(results)
        
        return results, summary
    
    def _generate_summary(self, results: List[ValidationResult]) -> ValidationSummary:
        """生成验证摘要"""
        total_plans = len(results)
        matched_plans = sum(1 for r in results if r.match_found)
        unmatched_plans = total_plans - matched_plans
        match_rate = matched_plans / total_plans if total_plans > 0 else 0.0
        
        # 计算平均相似度
        similarities = [r.similarity_score for r in results if r.similarity_score > 0]
        avg_similarity = sum(similarities) / len(similarities) if similarities else 0.0
        
        # 计算平均参数差异
        voltage_deltas = []
        current_deltas = []
        for r in results:
            voltage_delta = r.delta_params.get('voltage_V', 0.0)
            if voltage_delta not in [0.0, float('inf')] and voltage_delta is not None:
                voltage_deltas.append(abs(voltage_delta))
            current_delta = r.delta_params.get('current_density_A_dm2', 0.0)
            if current_delta not in [0.0, float('inf')] and current_delta is not None:
                current_deltas.append(abs(current_delta))
        
        avg_delta_voltage = sum(voltage_deltas) / len(voltage_deltas) if voltage_deltas else 0.0
        avg_delta_current = sum(current_deltas) / len(current_deltas) if current_deltas else 0.0
        
        # 统计引用频次
        source_counts = {}
        for r in results:
            for citation in r.nearest_citations:
                source = citation.get('source', 'unknown')
                source_counts[source] = source_counts.get(source, 0) + 1
        
        most_cited = sorted(source_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return ValidationSummary(
            total_plans=total_plans,
            matched_plans=matched_plans,
            unmatched_plans=unmatched_plans,
            match_rate=match_rate,
            avg_similarity=avg_similarity,
            avg_delta_voltage=avg_delta_voltage,
            avg_delta_current=avg_delta_current,
            most_cited_sources=most_cited,
            validation_time=datetime.now().isoformat()
        )
    
    def export_results(self, results: List[ValidationResult], summary: ValidationSummary, output_path: str) -> pathlib.Path:
        """导出验证结果到Excel"""
        output_path = pathlib.Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 准备数据
        matched_data = []
        unmatched_data = []
        
        for result in results:
            base_data = {
                'plan_id': result.plan_id,
                'system': result.system,
                'match_found': result.match_found,
                'similarity_score': result.similarity_score,
                'query_used': result.query_used,
                'validation_time': result.validation_time,
                'nearest_source_1': result.nearest_citations[0]['source'] if len(result.nearest_citations) > 0 else '',
                'nearest_page_1': result.nearest_citations[0]['page'] if len(result.nearest_citations) > 0 else '',
                'nearest_score_1': result.nearest_citations[0]['score'] if len(result.nearest_citations) > 0 else 0.0,
                'nearest_text_1': result.nearest_citations[0]['text'] if len(result.nearest_citations) > 0 else '',
                'delta_voltage_pct': result.delta_params.get('voltage_V', 0.0),
                'delta_current_pct': result.delta_params.get('current_density_A_dm2', 0.0),
                'delta_frequency_pct': result.delta_params.get('frequency_Hz', 0.0),
                'delta_duty_cycle_pct': result.delta_params.get('duty_cycle_pct', 0.0),
                'delta_time_pct': result.delta_params.get('time_min', 0.0)
            }
            
            # 添加更多最近邻引用
            for i in range(1, min(3, len(result.nearest_citations))):
                citation = result.nearest_citations[i]
                base_data[f'nearest_source_{i+1}'] = citation['source']
                base_data[f'nearest_page_{i+1}'] = citation['page']
                base_data[f'nearest_score_{i+1}'] = citation['score']
                base_data[f'nearest_text_{i+1}'] = citation['text']
            
            if result.match_found:
                matched_data.append(base_data)
            else:
                unmatched_data.append(base_data)
        
        # 创建Excel文件
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建摘要工作表
            summary_ws = wb.create_sheet("Summary")
            summary_ws.append(["验证摘要", ""])
            summary_ws.append(["总方案数", summary.total_plans])
            summary_ws.append(["匹配方案数", summary.matched_plans])
            summary_ws.append(["未匹配方案数", summary.unmatched_plans])
            summary_ws.append(["匹配率", f"{summary.match_rate*100:.1f}%"])
            summary_ws.append(["平均相似度", f"{summary.avg_similarity:.3f}"])
            summary_ws.append(["平均电压差异", f"{summary.avg_delta_voltage:.1f}%"])
            summary_ws.append(["平均电流差异", f"{summary.avg_delta_current:.1f}%"])
            summary_ws.append(["", ""])
            summary_ws.append(["最常引用文献", ""])
            for source, count in summary.most_cited_sources:
                summary_ws.append([source, count])
            
            # 设置摘要工作表样式
            for cell in summary_ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 创建匹配方案工作表
            if matched_data:
                matched_ws = wb.create_sheet("Matched")
                matched_df = pd.DataFrame(matched_data)
                
                # 写入表头
                for c_idx, col_name in enumerate(matched_df.columns, 1):
                    cell = matched_ws.cell(row=1, column=c_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                
                # 写入数据
                for r_idx, row in matched_df.iterrows():
                    for c_idx, value in enumerate(row, 1):
                        matched_ws.cell(row=r_idx+2, column=c_idx, value=value)
            
            # 创建未匹配方案工作表
            if unmatched_data:
                unmatched_ws = wb.create_sheet("Unmatched")
                unmatched_df = pd.DataFrame(unmatched_data)
                
                # 写入表头
                for c_idx, col_name in enumerate(unmatched_df.columns, 1):
                    cell = unmatched_ws.cell(row=1, column=c_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                # 写入数据
                for r_idx, row in unmatched_df.iterrows():
                    for c_idx, value in enumerate(row, 1):
                        unmatched_ws.cell(row=r_idx+2, column=c_idx, value=value)
            
            # 保存Excel文件
            wb.save(output_path)
            logger.info(f"Validation report exported to: {output_path}")
            
        except ImportError:
            # 如果没有openpyxl，降级到CSV
            logger.warning("openpyxl not available, exporting to CSV format")
            
            # 导出为CSV
            csv_path = output_path.with_suffix('.csv')
            all_data = matched_data + unmatched_data
            if all_data:
                df = pd.DataFrame(all_data)
                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                logger.info(f"Validation report exported to: {csv_path}")
                return csv_path
        
        return output_path

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="推荐验证脚本 - 将批量方案与KB文献进行对照验证",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  # 基本验证
  python scripts/validate_recommendations.py --plans tasks/batch_20250812_2246/plans.csv --kb datasets/index_store
  
  # 自定义TopK和阈值
  python scripts/validate_recommendations.py --plans tasks/batch_*/plans.csv --kb datasets/index_store --topk 5 --threshold 0.7
  
  # 指定输出路径
  python scripts/validate_recommendations.py --plans tasks/batch_20250812_2246/plans.csv --kb datasets/index_store --output validation_results.xlsx
        """
    )
    
    parser.add_argument("--plans", 
                       required=True,
                       help="批量方案CSV文件路径")
    
    parser.add_argument("--kb", 
                       required=True,
                       help="知识库索引目录路径")
    
    parser.add_argument("--topk", 
                       type=int, 
                       default=3,
                       help="返回最相近的K个文献片段 (默认: 3)")
    
    parser.add_argument("--threshold", 
                       type=float, 
                       default=0.6,
                       help="相似度匹配阈值 (默认: 0.6)")
    
    parser.add_argument("--output", 
                       type=str,
                       help="输出文件路径 (默认: 在plans同目录下生成validation_report.xlsx)")
    
    args = parser.parse_args()
    
    try:
        # 确定输出路径
        if args.output:
            output_path = args.output
        else:
            plans_path = pathlib.Path(args.plans)
            output_path = plans_path.parent / "validation_report.xlsx"
        
        # 创建验证器
        validator = RecommendationValidator(
            kb_path=args.kb,
            similarity_threshold=args.threshold
        )
        
        # 执行验证
        print(f"🔍 开始验证批量方案...")
        print(f"📁 方案文件: {args.plans}")
        print(f"📚 知识库: {args.kb}")
        print(f"🎯 Top-K: {args.topk}")
        print(f"📊 相似度阈值: {args.threshold}")
        
        results, summary = validator.validate_batch(args.plans, topk=args.topk)
        
        # 导出结果
        output_file = validator.export_results(results, summary, output_path)
        
        # 打印摘要
        print(f"\n📋 验证完成!")
        print(f"📊 验证摘要:")
        print(f"   - 总方案数: {summary.total_plans}")
        print(f"   - 匹配方案数: {summary.matched_plans}")
        print(f"   - 未匹配方案数: {summary.unmatched_plans}")
        print(f"   - 匹配率: {summary.match_rate*100:.1f}%")
        print(f"   - 平均相似度: {summary.avg_similarity:.3f}")
        print(f"   - 平均电压差异: {summary.avg_delta_voltage:.1f}%")
        print(f"   - 平均电流差异: {summary.avg_delta_current:.1f}%")
        
        if summary.most_cited_sources:
            print(f"\n📚 最常引用文献:")
            for source, count in summary.most_cited_sources[:3]:
                print(f"   - {source}: {count}次")
        
        print(f"\n📄 详细报告: {output_file}")
        
        # 根据匹配率给出建议
        if summary.match_rate >= 0.8:
            print(f"\n✅ 验证结果良好！大部分方案都有历史先例支撑。")
        elif summary.match_rate >= 0.5:
            print(f"\n⚠️ 验证结果一般，建议检查未匹配方案的创新性和可行性。")
        else:
            print(f"\n❌ 验证结果较差，多数方案缺乏文献支撑，建议重新评估。")
        
    except Exception as e:
        logger.error(f"Validation failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
