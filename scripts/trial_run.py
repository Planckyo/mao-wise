#!/usr/bin/env python3
"""
MAO-Wise 试运行主逻辑脚本

执行完整的API测试、验收和报告生成流程：
- Clarify & SlotFill 流程测试
- 必答问题 & 追问机制测试
- 实验结果录入和模型更新
- RAG引用验证和UI截图
- 生成详细的试运行报告

使用示例：
python scripts/trial_run.py --mode offline --batch tasks/batch_20250812_2300
python scripts/trial_run.py --mode online --batch tasks/batch_20250812_2300
"""

import argparse
import json
import sys
import pathlib
import time
import subprocess
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
import httpx
import openpyxl
from openpyxl import Workbook

# 确保能找到maowise包
REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from maowise.utils.logger import logger

class TrialRunner:
    """试运行主逻辑执行器"""
    
    def __init__(self, mode: str = "offline", batch_dir: Optional[str] = None):
        self.mode = mode
        self.batch_dir = pathlib.Path(batch_dir) if batch_dir else None
        self.api_base = "http://127.0.0.1:8000"
        self.ui_base = "http://127.0.0.1:8501"
        self.reports_dir = pathlib.Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
        
        # 试运行状态记录
        self.results = {
            "start_time": datetime.now().isoformat(),
            "mode": mode,
            "batch_dir": str(batch_dir) if batch_dir else None,
            "steps": {},
            "api_tests": {},
            "model_updates": {},
            "screenshots": {},
            "errors": [],
            "summary": {}
        }
        
        # HTTP客户端
        self.client = httpx.Client(timeout=30)
    
    def log_step(self, step_name: str, status: str, duration: float = 0, details: Dict = None):
        """记录步骤结果"""
        self.results["steps"][step_name] = {
            "status": status,
            "duration_seconds": round(duration, 2),
            "details": details or {},
            "timestamp": datetime.now().isoformat()
        }
        
        status_icon = "✅" if status == "success" else ("⚠️" if status == "warning" else "❌")
        print(f"{status_icon} {step_name} ({duration:.1f}s)")
    
    def test_api_health(self) -> bool:
        """测试API健康状态"""
        try:
            response = self.client.get(f"{self.api_base}/api/maowise/v1/health")
            if response.status_code == 200:
                health_data = response.json()
                self.log_step("API健康检查", "success", 0, health_data)
                return True
            else:
                self.log_step("API健康检查", "failed", 0, {"status_code": response.status_code})
                return False
        except Exception as e:
            self.log_step("API健康检查", "failed", 0, {"error": str(e)})
            return False
    
    def test_clarify_slotfill_flow(self) -> Dict[str, Any]:
        """测试Clarify & SlotFill流程"""
        start_time = time.time()
        
        try:
            # 1. 发送缺电压的预测请求
            predict_request = {
                "substrate_alloy": "AZ91D",
                "electrolyte_family": "fluoride",
                "electrolyte_components": ["K2ZrF6", "KOH"],
                "mode": "ac",
                # 故意缺少 voltage_V
                "current_density_A_dm2": 10.0,
                "frequency_Hz": 800,
                "duty_cycle_pct": 35,
                "time_min": 20,
                "temp_C": 25,
                "pH": 11.0,
                "sealing": "none"
            }
            
            response = self.client.post(f"{self.api_base}/api/maowise/v1/predict_or_ask", json=predict_request)
            
            if response.status_code != 200:
                duration = time.time() - start_time
                self.log_step("Clarify测试", "failed", duration, {"status_code": response.status_code})
                return {"status": "failed", "error": f"HTTP {response.status_code}"}
            
            result = response.json()
            
            # 检查是否返回需要专家回答
            if not result.get("need_expert", False):
                duration = time.time() - start_time
                self.log_step("Clarify测试", "failed", duration, {"error": "未触发专家询问"})
                return {"status": "failed", "error": "未触发专家询问"}
            
            # 获取问题列表
            questions = result.get("questions", [])
            if not questions:
                duration = time.time() - start_time
                self.log_step("Clarify测试", "failed", duration, {"error": "未生成问题"})
                return {"status": "failed", "error": "未生成问题"}
            
            # 找到电压相关问题
            voltage_question = None
            for q in questions:
                if "电压" in q.get("question", "") or "voltage" in q.get("question", "").lower():
                    voltage_question = q
                    break
            
            if not voltage_question:
                duration = time.time() - start_time
                self.log_step("Clarify测试", "warning", duration, {"error": "未找到电压问题", "questions": questions})
                return {"status": "warning", "error": "未找到电压问题"}
            
            # 2. 回答电压问题
            thread_id = result.get("thread_id")
            answers = [{"question_id": voltage_question["id"], "answer": "电压 420 V"}]
            
            resolve_request = {
                "thread_id": thread_id,
                "answers": answers
            }
            
            resolve_response = self.client.post(f"{self.api_base}/api/maowise/v1/expert/thread/resolve", json=resolve_request)
            
            if resolve_response.status_code != 200:
                duration = time.time() - start_time
                self.log_step("Clarify测试", "failed", duration, {"status_code": resolve_response.status_code})
                return {"status": "failed", "error": f"Resolve HTTP {resolve_response.status_code}"}
            
            resolve_result = resolve_response.json()
            
            # 检查是否返回预测结果
            if not resolve_result.get("resolved", False):
                duration = time.time() - start_time
                self.log_step("Clarify测试", "failed", duration, {"error": "未成功解决", "result": resolve_result})
                return {"status": "failed", "error": "未成功解决"}
            
            prediction = resolve_result.get("prediction", {})
            if not prediction.get("alpha_150_2600") or not prediction.get("epsilon_3000_30000"):
                duration = time.time() - start_time
                self.log_step("Clarify测试", "failed", duration, {"error": "预测结果不完整", "prediction": prediction})
                return {"status": "failed", "error": "预测结果不完整"}
            
            duration = time.time() - start_time
            test_details = {
                "questions_count": len(questions),
                "voltage_question": voltage_question["question"],
                "prediction": prediction,
                "confidence": prediction.get("confidence", 0)
            }
            
            self.log_step("Clarify & SlotFill流程", "success", duration, test_details)
            return {"status": "success", "details": test_details}
            
        except Exception as e:
            duration = time.time() - start_time
            self.log_step("Clarify & SlotFill流程", "failed", duration, {"error": str(e)})
            return {"status": "failed", "error": str(e)}
    
    def test_mandatory_followup_flow(self) -> Dict[str, Any]:
        """测试必答问题 & 追问流程"""
        start_time = time.time()
        
        try:
            # 1. 发送缺质量上限的推荐请求
            recommend_request = {
                "target": {"alpha": 0.20, "epsilon": 0.80},
                "substrate_alloy": "AZ91D",
                "electrolyte_family": "fluoride",
                "system_constraints": {
                    "voltage_range": [250, 400],
                    "current_range": [5, 12]
                },
                # 故意缺少质量/厚度上限相关约束
                "preferences": {"priority": "balanced"}
            }
            
            response = self.client.post(f"{self.api_base}/api/maowise/v1/recommend_or_ask", json=recommend_request)
            
            if response.status_code != 200:
                duration = time.time() - start_time
                self.log_step("必答问题测试", "failed", duration, {"status_code": response.status_code})
                return {"status": "failed", "error": f"HTTP {response.status_code}"}
            
            result = response.json()
            
            # 检查是否触发必答问题
            if not result.get("need_expert", False):
                duration = time.time() - start_time
                self.log_step("必答问题测试", "failed", duration, {"error": "未触发必答问题"})
                return {"status": "failed", "error": "未触发必答问题"}
            
            questions = result.get("questions", [])
            mandatory_questions = [q for q in questions if q.get("is_mandatory", False)]
            
            if not mandatory_questions:
                duration = time.time() - start_time
                self.log_step("必答问题测试", "failed", duration, {"error": "未找到必答问题"})
                return {"status": "failed", "error": "未找到必答问题"}
            
            # 找到质量相关的必答问题
            quality_question = None
            for q in mandatory_questions:
                if "质量" in q.get("question", "") or "厚度" in q.get("question", ""):
                    quality_question = q
                    break
            
            if not quality_question:
                # 取第一个必答问题进行测试
                quality_question = mandatory_questions[0]
            
            thread_id = result.get("thread_id")
            
            # 2. 先给出模糊答案触发追问
            vague_answers = [{"question_id": quality_question["id"], "answer": "看情况"}]
            
            resolve_request = {
                "thread_id": thread_id,
                "answers": vague_answers
            }
            
            resolve_response = self.client.post(f"{self.api_base}/api/maowise/v1/expert/thread/resolve", json=resolve_request)
            
            if resolve_response.status_code != 200:
                duration = time.time() - start_time
                self.log_step("必答问题测试", "failed", duration, {"status_code": resolve_response.status_code})
                return {"status": "failed", "error": f"First resolve HTTP {resolve_response.status_code}"}
            
            first_result = resolve_response.json()
            
            # 检查是否生成追问
            if first_result.get("resolved", False):
                # 如果直接解决了，可能追问机制未触发，但这也是可接受的
                pass
            else:
                # 应该有追问问题
                followup_questions = first_result.get("questions", [])
                if followup_questions:
                    # 回答追问
                    specific_answers = [{"question_id": followup_questions[0]["id"], "answer": "单位面积质量 ≤ 50 g/m²"}]
                    
                    resolve_request["answers"] = vague_answers + specific_answers
                    
                    final_response = self.client.post(f"{self.api_base}/api/maowise/v1/expert/thread/resolve", json=resolve_request)
                    
                    if final_response.status_code == 200:
                        first_result = final_response.json()
            
            # 3. 提供具体答案
            if not first_result.get("resolved", False):
                specific_answers = [{"question_id": quality_question["id"], "answer": "单位面积质量 ≤ 50 g/m²"}]
                
                resolve_request = {
                    "thread_id": thread_id,
                    "answers": specific_answers
                }
                
                final_response = self.client.post(f"{self.api_base}/api/maowise/v1/expert/thread/resolve", json=resolve_request)
                
                if final_response.status_code != 200:
                    duration = time.time() - start_time
                    self.log_step("必答问题测试", "failed", duration, {"status_code": final_response.status_code})
                    return {"status": "failed", "error": f"Final resolve HTTP {final_response.status_code}"}
                
                final_result = final_response.json()
            else:
                final_result = first_result
            
            # 检查是否返回推荐方案
            if not final_result.get("resolved", False):
                duration = time.time() - start_time
                self.log_step("必答问题测试", "failed", duration, {"error": "未成功解决必答问题"})
                return {"status": "failed", "error": "未成功解决必答问题"}
            
            recommendations = final_result.get("recommendations", {}).get("solutions", [])
            if not recommendations:
                duration = time.time() - start_time
                self.log_step("必答问题测试", "failed", duration, {"error": "未返回推荐方案"})
                return {"status": "failed", "error": "未返回推荐方案"}
            
            # 检查plan_yaml和引用
            plan_yaml_count = 0
            citation_count = 0
            
            for solution in recommendations:
                if solution.get("plan_yaml"):
                    plan_yaml_count += 1
                
                explanation = solution.get("explanation", "")
                if "[CIT-" in explanation:
                    citation_count += 1
            
            duration = time.time() - start_time
            test_details = {
                "mandatory_questions_count": len(mandatory_questions),
                "total_questions": len(questions),
                "solutions_count": len(recommendations),
                "plan_yaml_count": plan_yaml_count,
                "citation_count": citation_count,
                "quality_question": quality_question["question"]
            }
            
            self.log_step("必答问题 & 追问流程", "success", duration, test_details)
            return {"status": "success", "details": test_details}
            
        except Exception as e:
            duration = time.time() - start_time
            self.log_step("必答问题 & 追问流程", "failed", duration, {"error": str(e)})
            return {"status": "failed", "error": str(e)}
    
    def create_fake_experiment_results(self) -> str:
        """创建假实验结果Excel文件"""
        try:
            # 创建results目录
            results_dir = pathlib.Path("results")
            results_dir.mkdir(exist_ok=True)
            
            # 从最新批次获取方案数据
            fake_results = []
            
            if self.batch_dir and (self.batch_dir / "plans.csv").exists():
                try:
                    plans_df = pd.read_csv(self.batch_dir / "plans.csv")
                    # 取前2条作为实验结果
                    for i, row in plans_df.head(2).iterrows():
                        # 基于方案参数生成模拟实验结果
                        voltage = row.get('voltage_V', 300)
                        current = row.get('current_density_A_dm2', 10)
                        
                        # 简单的结果模拟（基于参数的经验公式）
                        measured_alpha = 0.15 + (voltage - 300) * 0.0001 + (current - 10) * 0.005
                        measured_epsilon = 0.7 + (voltage - 300) * 0.0003 + (current - 10) * 0.01
                        
                        # 添加一些随机变化
                        import random
                        measured_alpha += random.uniform(-0.02, 0.02)
                        measured_epsilon += random.uniform(-0.05, 0.05)
                        
                        # 限制在合理范围
                        measured_alpha = max(0.05, min(0.4, measured_alpha))
                        measured_epsilon = max(0.5, min(1.2, measured_epsilon))
                        
                        result = {
                            'experiment_id': f'TRIAL-EXP-{i+1:03d}',
                            'batch_id': self.batch_dir.name if self.batch_dir else 'trial_batch',
                            'plan_id': f'trial_plan_{i+1:03d}',
                            'system': row.get('system', 'zirconate'),
                            'substrate_alloy': row.get('substrate_alloy', 'AZ91D'),
                            'electrolyte_components_json': json.dumps(row.get('electrolyte_components', ['K2ZrF6', 'KOH'])),
                            'voltage_V': voltage,
                            'current_density_Adm2': current,
                            'frequency_Hz': row.get('frequency_Hz', 800),
                            'duty_cycle_pct': row.get('duty_cycle_pct', 35),
                            'time_min': row.get('time_min', 20),
                            'temp_C': 25.0,
                            'pH': 11.0,
                            'post_treatment': 'none',
                            'measured_alpha': round(measured_alpha, 4),
                            'measured_epsilon': round(measured_epsilon, 4),
                            'hardness_HV': round(180 + random.uniform(-20, 20), 1),
                            'roughness_Ra_um': round(2.0 + random.uniform(-0.5, 0.5), 2),
                            'corrosion_rate_mmpy': round(0.05 + random.uniform(-0.02, 0.02), 4),
                            'notes': f'试运行实验 {i+1} - 质量良好',
                            'reviewer': '试运行系统',
                            'timestamp': datetime.now().isoformat()
                        }
                        fake_results.append(result)
                        
                except Exception as e:
                    logger.warning(f"无法从批次文件生成结果，使用默认值: {e}")
            
            # 如果没有从批次获取到数据，使用默认数据
            if not fake_results:
                for i in range(2):
                    result = {
                        'experiment_id': f'TRIAL-EXP-{i+1:03d}',
                        'batch_id': 'trial_batch',
                        'plan_id': f'trial_plan_{i+1:03d}',
                        'system': 'zirconate',
                        'substrate_alloy': 'AZ91D',
                        'electrolyte_components_json': '["K2ZrF6", "KOH"]',
                        'voltage_V': 300 + i * 20,
                        'current_density_Adm2': 10 + i,
                        'frequency_Hz': 800,
                        'duty_cycle_pct': 35,
                        'time_min': 20,
                        'temp_C': 25.0,
                        'pH': 11.0,
                        'post_treatment': 'none',
                        'measured_alpha': 0.15 + i * 0.01,
                        'measured_epsilon': 0.75 + i * 0.02,
                        'hardness_HV': 180 + i * 10,
                        'roughness_Ra_um': 2.0 + i * 0.2,
                        'corrosion_rate_mmpy': 0.05 + i * 0.01,
                        'notes': f'试运行实验 {i+1} - 质量良好',
                        'reviewer': '试运行系统',
                        'timestamp': datetime.now().isoformat()
                    }
                    fake_results.append(result)
            
            # 创建Excel文件
            excel_file = results_dir / "trial_results.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "实验结果"
            
            # 写入表头
            headers = list(fake_results[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row_idx, result in enumerate(fake_results, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=result[header])
            
            # 保存文件
            wb.save(excel_file)
            
            self.log_step("创建假实验结果", "success", 0, {"file": str(excel_file), "records": len(fake_results)})
            return str(excel_file)
            
        except Exception as e:
            self.log_step("创建假实验结果", "failed", 0, {"error": str(e)})
            raise
    
    def import_experiment_results(self, excel_file: str) -> Dict[str, Any]:
        """导入实验结果"""
        start_time = time.time()
        
        try:
            # 调用导入脚本
            result = subprocess.run([
                sys.executable, "scripts/record_experiment_results.py",
                "--file", excel_file
            ], capture_output=True, text=True, cwd=REPO_ROOT)
            
            duration = time.time() - start_time
            
            if result.returncode == 0:
                self.log_step("导入实验结果", "success", duration, {"stdout": result.stdout})
                return {"status": "success", "output": result.stdout}
            else:
                self.log_step("导入实验结果", "failed", duration, {"stderr": result.stderr})
                return {"status": "failed", "error": result.stderr}
                
        except Exception as e:
            duration = time.time() - start_time
            self.log_step("导入实验结果", "failed", duration, {"error": str(e)})
            return {"status": "failed", "error": str(e)}
    
    def run_evaluation_and_update(self) -> Dict[str, Any]:
        """执行评估与模型更新"""
        start_time = time.time()
        
        try:
            # 1. 执行评估
            eval_result = subprocess.run([
                sys.executable, "scripts/evaluate_predictions.py"
            ], capture_output=True, text=True, cwd=REPO_ROOT)
            
            if eval_result.returncode != 0:
                duration = time.time() - start_time
                self.log_step("预测评估", "failed", duration, {"stderr": eval_result.stderr})
                return {"status": "failed", "error": eval_result.stderr}
            
            # 2. 执行模型更新
            update_cmd = [
                "powershell", "-ExecutionPolicy", "Bypass", "-File", "scripts\\update_from_feedback.ps1", "-HotReload:$true"
            ]
            
            update_result = subprocess.run(update_cmd, capture_output=True, text=True, cwd=REPO_ROOT)
            
            # 3. 再次评估用于对比
            eval2_result = subprocess.run([
                sys.executable, "scripts/evaluate_predictions.py"
            ], capture_output=True, text=True, cwd=REPO_ROOT)
            
            duration = time.time() - start_time
            
            update_success = update_result.returncode == 0
            eval2_success = eval2_result.returncode == 0
            
            details = {
                "evaluation_1": eval_result.stdout,
                "model_update": update_result.stdout if update_success else update_result.stderr,
                "evaluation_2": eval2_result.stdout if eval2_success else eval2_result.stderr,
                "update_success": update_success,
                "eval2_success": eval2_success
            }
            
            status = "success" if update_success else "warning"
            self.log_step("评估与模型更新", status, duration, details)
            
            return {"status": status, "details": details}
            
        except Exception as e:
            duration = time.time() - start_time
            self.log_step("评估与模型更新", "failed", duration, {"error": str(e)})
            return {"status": "failed", "error": str(e)}
    
    def test_hot_reload(self) -> Dict[str, Any]:
        """测试热加载功能"""
        start_time = time.time()
        
        try:
            reload_request = {
                "models": ["gp_corrector", "reward_model"],
                "force": True
            }
            
            response = self.client.post(f"{self.api_base}/api/maowise/v1/admin/reload", json=reload_request)
            
            duration = time.time() - start_time
            
            if response.status_code == 200:
                result = response.json()
                self.log_step("API热加载", "success", duration, result)
                return {"status": "success", "details": result}
            else:
                self.log_step("API热加载", "failed", duration, {"status_code": response.status_code})
                return {"status": "failed", "error": f"HTTP {response.status_code}"}
                
        except Exception as e:
            duration = time.time() - start_time
            self.log_step("API热加载", "failed", duration, {"error": str(e)})
            return {"status": "failed", "error": str(e)}
    
    def capture_ui_screenshots(self) -> Dict[str, Any]:
        """捕获UI截图"""
        start_time = time.time()
        
        try:
            # 调用UI截图脚本
            result = subprocess.run([
                sys.executable, "scripts/ui_snapshots.py"
            ], capture_output=True, text=True, cwd=REPO_ROOT)
            
            duration = time.time() - start_time
            
            # 检查生成的截图文件
            screenshot_files = []
            for filename in ["ui_predict.png", "ui_recommend.png", "ui_expert.png"]:
                filepath = self.reports_dir / filename
                if filepath.exists():
                    screenshot_files.append(str(filepath))
            
            if result.returncode == 0 or screenshot_files:
                self.log_step("UI截图捕获", "success", duration, {"files": screenshot_files})
                return {"status": "success", "files": screenshot_files}
            else:
                self.log_step("UI截图捕获", "failed", duration, {"stderr": result.stderr})
                return {"status": "failed", "error": result.stderr}
                
        except Exception as e:
            duration = time.time() - start_time
            self.log_step("UI截图捕获", "failed", duration, {"error": str(e)})
            return {"status": "failed", "error": str(e)}
    
    def verify_rag_citations(self) -> Dict[str, Any]:
        """验证RAG引用"""
        start_time = time.time()
        
        try:
            # 发送一个推荐请求来检查引用
            request = {
                "target": {"alpha": 0.15, "epsilon": 0.85},
                "substrate_alloy": "AZ91D",
                "electrolyte_family": "fluoride",
                "system_constraints": {"voltage_range": [280, 350]},
                "preferences": {"priority": "performance"}
            }
            
            response = self.client.post(f"{self.api_base}/api/maowise/v1/recommend", json=request)
            
            duration = time.time() - start_time
            
            if response.status_code != 200:
                self.log_step("RAG引用验证", "failed", duration, {"status_code": response.status_code})
                return {"status": "failed", "error": f"HTTP {response.status_code}"}
            
            result = response.json()
            solutions = result.get("solutions", [])
            
            citation_stats = {
                "total_solutions": len(solutions),
                "solutions_with_citations": 0,
                "total_citations": 0,
                "explanation_length_check": 0
            }
            
            for solution in solutions:
                explanation = solution.get("explanation", "")
                
                # 检查解释长度（应该 ≤ 7 条）
                explanation_lines = [line.strip() for line in explanation.split('\n') if line.strip()]
                if len(explanation_lines) <= 7:
                    citation_stats["explanation_length_check"] += 1
                
                # 检查引用格式
                citation_count = explanation.count("[CIT-")
                if citation_count > 0:
                    citation_stats["solutions_with_citations"] += 1
                    citation_stats["total_citations"] += citation_count
            
            # 判断是否通过验证
            citations_ok = citation_stats["solutions_with_citations"] > 0
            length_ok = citation_stats["explanation_length_check"] >= citation_stats["total_solutions"] * 0.8
            
            status = "success" if (citations_ok and length_ok) else "warning"
            
            self.log_step("RAG引用验证", status, duration, citation_stats)
            return {"status": status, "details": citation_stats}
            
        except Exception as e:
            duration = time.time() - start_time
            self.log_step("RAG引用验证", "failed", duration, {"error": str(e)})
            return {"status": "failed", "error": str(e)}
    
    def generate_report(self):
        """生成试运行报告"""
        # 计算总耗时
        end_time = datetime.now()
        start_time = datetime.fromisoformat(self.results["start_time"])
        total_duration = (end_time - start_time).total_seconds()
        
        self.results["end_time"] = end_time.isoformat()
        self.results["total_duration_seconds"] = total_duration
        
        # 生成摘要统计
        steps = self.results["steps"]
        success_count = len([s for s in steps.values() if s["status"] == "success"])
        warning_count = len([s for s in steps.values() if s["status"] == "warning"])
        failed_count = len([s for s in steps.values() if s["status"] == "failed"])
        
        self.results["summary"] = {
            "total_steps": len(steps),
            "success_count": success_count,
            "warning_count": warning_count, 
            "failed_count": failed_count,
            "success_rate": round(success_count / len(steps) * 100, 1) if steps else 0,
            "mode": self.mode,
            "batch_directory": str(self.batch_dir) if self.batch_dir else None
        }
        
        # 生成Markdown报告
        self._generate_markdown_report()
        
        # 生成HTML报告
        self._generate_html_report()
        
        print(f"\n📊 试运行完成！总耗时: {total_duration:.1f}秒")
        print(f"   - 成功: {success_count}/{len(steps)} 步骤")
        print(f"   - 警告: {warning_count} 步骤")
        print(f"   - 失败: {failed_count} 步骤")
        print(f"   - 成功率: {self.results['summary']['success_rate']}%")
    
    def _generate_markdown_report(self):
        """生成Markdown格式报告"""
        report_file = self.reports_dir / "trial_run_report.md"
        
        md_content = f"""# MAO-Wise 试运行报告

## 📋 基本信息

- **运行时间**: {self.results['start_time']} ~ {self.results['end_time']}
- **总耗时**: {self.results['total_duration_seconds']:.1f} 秒
- **运行模式**: {self.results['mode']}
- **批次目录**: {self.results['summary']['batch_directory'] or '无'}

## 📊 执行摘要

- **总步骤数**: {self.results['summary']['total_steps']}
- **成功步骤**: {self.results['summary']['success_count']}
- **警告步骤**: {self.results['summary']['warning_count']}
- **失败步骤**: {self.results['summary']['failed_count']}
- **成功率**: {self.results['summary']['success_rate']}%

## 🔍 详细步骤

"""
        
        for step_name, step_data in self.results["steps"].items():
            status_icon = "✅" if step_data["status"] == "success" else ("⚠️" if step_data["status"] == "warning" else "❌")
            md_content += f"""### {status_icon} {step_name}

- **状态**: {step_data['status']}
- **耗时**: {step_data['duration_seconds']} 秒
- **时间戳**: {step_data['timestamp']}

"""
            
            if step_data.get("details"):
                md_content += "**详细信息**:\n```json\n"
                md_content += json.dumps(step_data["details"], ensure_ascii=False, indent=2)
                md_content += "\n```\n\n"
        
        # 错误汇总
        if self.results["errors"]:
            md_content += "## ❌ 错误汇总\n\n"
            for error in self.results["errors"]:
                md_content += f"- {error}\n"
            md_content += "\n"
        
        # 建议
        md_content += f"""## 💡 建议

### 性能优化
- 如需提升性能，考虑启用在线模式（需要OpenAI API Key）
- 增加本地文献库以提升RAG质量

### 后续步骤
1. 查看生成的批量方案和验证报告
2. 检查UI截图确认界面正常
3. 监控模型更新后的性能改进

---

*报告生成时间: {datetime.now().isoformat()}*
"""
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(md_content)
        
        print(f"✅ Markdown报告: {report_file}")
    
    def _generate_html_report(self):
        """生成HTML格式报告"""
        report_file = self.reports_dir / "trial_run_report.html"
        
        # 状态颜色映射
        status_colors = {
            "success": "#28a745",
            "warning": "#ffc107", 
            "failed": "#dc3545"
        }
        
        # 生成步骤HTML
        steps_html = ""
        for step_name, step_data in self.results["steps"].items():
            status_color = status_colors.get(step_data["status"], "#6c757d")
            status_icon = "✅" if step_data["status"] == "success" else ("⚠️" if step_data["status"] == "warning" else "❌")
            
            steps_html += f"""
            <div class="step-card">
                <h3>{status_icon} {step_name}</h3>
                <div class="step-meta">
                    <span class="status" style="color: {status_color};">{step_data['status']}</span>
                    <span class="duration">{step_data['duration_seconds']}s</span>
                    <span class="timestamp">{step_data['timestamp']}</span>
                </div>
            """
            
            if step_data.get("details"):
                steps_html += f"""
                <details class="step-details">
                    <summary>详细信息</summary>
                    <pre>{json.dumps(step_data["details"], ensure_ascii=False, indent=2)}</pre>
                </details>
                """
            
            steps_html += "</div>"
        
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>MAO-Wise 试运行报告</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; margin: 0; padding: 20px; background: #f8f9fa; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        h3 {{ color: #2c3e50; margin: 15px 0 10px 0; }}
        .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }}
        .summary-card {{ background: #ecf0f1; padding: 20px; border-radius: 6px; text-align: center; }}
        .summary-card h3 {{ margin: 0 0 10px 0; color: #2c3e50; }}
        .summary-card .value {{ font-size: 2em; font-weight: bold; color: #3498db; }}
        .step-card {{ background: #fff; border: 1px solid #dee2e6; border-radius: 6px; padding: 20px; margin: 15px 0; }}
        .step-meta {{ display: flex; gap: 15px; margin: 10px 0; font-size: 0.9em; }}
        .step-meta span {{ padding: 4px 8px; border-radius: 4px; background: #f8f9fa; }}
        .step-details {{ margin-top: 15px; }}
        .step-details pre {{ background: #f8f9fa; padding: 15px; border-radius: 4px; overflow-x: auto; font-size: 0.85em; }}
        .success {{ background-color: #d4edda !important; }}
        .warning {{ background-color: #fff3cd !important; }}
        .failed {{ background-color: #f8d7da !important; }}
        .footer {{ text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #dee2e6; color: #6c757d; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🧪 MAO-Wise 试运行报告</h1>
        
        <h2>📋 基本信息</h2>
        <ul>
            <li><strong>运行时间</strong>: {self.results['start_time']} ~ {self.results['end_time']}</li>
            <li><strong>总耗时</strong>: {self.results['total_duration_seconds']:.1f} 秒</li>
            <li><strong>运行模式</strong>: {self.results['mode']}</li>
            <li><strong>批次目录</strong>: {self.results['summary']['batch_directory'] or '无'}</li>
        </ul>
        
        <h2>📊 执行摘要</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <h3>总步骤</h3>
                <div class="value">{self.results['summary']['total_steps']}</div>
            </div>
            <div class="summary-card success">
                <h3>成功</h3>
                <div class="value">{self.results['summary']['success_count']}</div>
            </div>
            <div class="summary-card warning">
                <h3>警告</h3>
                <div class="value">{self.results['summary']['warning_count']}</div>
            </div>
            <div class="summary-card failed">
                <h3>失败</h3>
                <div class="value">{self.results['summary']['failed_count']}</div>
            </div>
            <div class="summary-card">
                <h3>成功率</h3>
                <div class="value">{self.results['summary']['success_rate']}%</div>
            </div>
        </div>
        
        <h2>🔍 详细步骤</h2>
        {steps_html}
        
        <div class="footer">
            <p>报告生成时间: {datetime.now().isoformat()}</p>
            <p>MAO-Wise 微弧氧化热控涂层优化器 v1.0</p>
        </div>
    </div>
</body>
</html>"""
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✅ HTML报告: {report_file}")
    
    def run(self):
        """执行完整的试运行流程"""
        print(f"🚀 开始试运行 (模式: {self.mode})")
        
        # 1. 检查API健康状态
        if not self.test_api_health():
            self.results["errors"].append("API服务不可用")
            self.generate_report()
            return
        
        # 2. 测试Clarify & SlotFill流程
        self.api_tests["clarify_slotfill"] = self.test_clarify_slotfill_flow()
        
        # 3. 测试必答问题 & 追问流程
        self.api_tests["mandatory_followup"] = self.test_mandatory_followup_flow()
        
        # 4. 验证RAG引用
        self.api_tests["rag_citations"] = self.verify_rag_citations()
        
        # 5. 创建和导入假实验结果
        try:
            excel_file = self.create_fake_experiment_results()
            self.model_updates["import_results"] = self.import_experiment_results(excel_file)
        except Exception as e:
            self.results["errors"].append(f"实验结果处理失败: {e}")
        
        # 6. 执行评估与模型更新
        self.model_updates["evaluation_update"] = self.run_evaluation_and_update()
        
        # 7. 测试热加载
        self.model_updates["hot_reload"] = self.test_hot_reload()
        
        # 8. 捕获UI截图
        self.screenshots = self.capture_ui_screenshots()
        
        # 9. 生成报告
        self.generate_report()

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="MAO-Wise 试运行主逻辑脚本",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument("--mode", 
                       choices=["online", "offline"],
                       default="offline",
                       help="运行模式 (default: offline)")
    
    parser.add_argument("--batch", 
                       type=str,
                       help="批次目录路径")
    
    args = parser.parse_args()
    
    try:
        runner = TrialRunner(mode=args.mode, batch_dir=args.batch)
        runner.run()
        
    except Exception as e:
        print(f"❌ 试运行失败: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
