#!/usr/bin/env python3
"""
测试推荐验证功能

验证功能：
- 使用微型corpus进行验证测试
- 断言能生成Excel报告
- 验证match_found字段存在
- 测试参数差异计算
- 验证摘要统计正确性
"""

import pytest
import csv
import json
import tempfile
import shutil
import pathlib
import sys
from unittest.mock import Mock, patch, MagicMock
import pandas as pd

# 确保能找到maowise包
REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.validate_recommendations import RecommendationValidator, ValidationResult, ValidationSummary

class TestRecommendationValidator:
    """推荐验证器测试"""
    
    @pytest.fixture
    def temp_dir(self):
        """临时目录fixture"""
        temp_dir = tempfile.mkdtemp()
        yield pathlib.Path(temp_dir)
        shutil.rmtree(temp_dir)
    
    @pytest.fixture
    def mock_kb_path(self, temp_dir):
        """模拟知识库路径"""
        kb_path = temp_dir / "index_store"
        kb_path.mkdir(parents=True)
        return kb_path
    
    @pytest.fixture
    def sample_plans_csv(self, temp_dir):
        """创建测试用的方案CSV文件"""
        plans_data = [
            {
                'plan_id': 'test_plan_001',
                'system': 'silicate',
                'voltage': '300',
                'current_density': '10.0',
                'frequency': '500',
                'duty_cycle': '30',
                'time': '15',
                'alpha': '0.20',
                'epsilon': '0.80',
                'confidence': '0.75'
            },
            {
                'plan_id': 'test_plan_002',
                'system': 'zirconate',
                'voltage': '250',
                'current_density': '8.0',
                'frequency': '800',
                'duty_cycle': '25',
                'time': '12',
                'alpha': '0.18',
                'epsilon': '0.85',
                'confidence': '0.68'
            }
        ]
        
        csv_path = temp_dir / "test_plans.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=plans_data[0].keys())
            writer.writeheader()
            writer.writerows(plans_data)
        
        return csv_path
    
    @pytest.fixture
    def mock_search_results(self):
        """模拟搜索结果"""
        return {
            'results': [
                {
                    'source': 'test_paper_1.pdf',
                    'page': 3,
                    'score': 0.85,
                    'text': 'AZ91 substrate with Na2SiO3 10g/L electrolyte, processed at 320V, 12A/dm2, 500Hz frequency, 30% duty cycle for 15min.'
                },
                {
                    'source': 'test_paper_2.pdf',
                    'page': 7,
                    'score': 0.72,
                    'text': 'Silicate-based MAO coating with voltage 300V and current density 10A/dm2, resulting in good thermal properties.'
                },
                {
                    'source': 'test_paper_3.pdf',
                    'page': 2,
                    'score': 0.65,
                    'text': 'Micro arc oxidation process using alkaline electrolyte, 280V voltage, frequency 450Hz.'
                }
            ]
        }
    
    def test_construct_query(self, mock_kb_path):
        """测试查询构造"""
        validator = RecommendationValidator(str(mock_kb_path))
        
        plan = {
            'system': 'silicate',
            'voltage': 300,
            'current_density': 10.0,
            'frequency': 500,
            'duty_cycle': 30,
            'time': 15
        }
        
        query = validator._construct_query(plan)
        
        assert 'silicate' in query.lower()
        assert 'na2sio3' in query.lower()
        assert '300v' in query.lower()
        assert '10.0a/dm2' in query.lower()
    
    def test_extract_plan_params(self, mock_kb_path):
        """测试方案参数提取"""
        validator = RecommendationValidator(str(mock_kb_path))
        
        plan = {
            'voltage_V': '300',
            'current_density_A_dm2': '10.5',
            'frequency_Hz': '500',
            'duty_cycle_pct': '30',
            'time_min': '15'
        }
        
        params = validator._extract_plan_params(plan)
        
        assert params['voltage_V'] == 300.0
        assert params['current_density_A_dm2'] == 10.5
        assert params['frequency_Hz'] == 500.0
        assert params['duty_cycle_pct'] == 30.0
        assert params['time_min'] == 15.0
    
    def test_extract_citation_params(self, mock_kb_path):
        """测试文献参数提取"""
        validator = RecommendationValidator(str(mock_kb_path))
        
        citation_text = "AZ91 substrate processed at 320V, 12A/dm2, 500Hz frequency, 30% duty cycle for 15min."
        
        params = validator._extract_citation_params(citation_text)
        
        assert params['voltage_V'] == 320.0
        assert params['current_density_A_dm2'] == 12.0
        assert params['frequency_Hz'] == 500.0
        assert params['duty_cycle_pct'] == 30.0
        assert params['time_min'] == 15.0
    
    def test_calculate_param_delta(self, mock_kb_path):
        """测试参数差异计算"""
        validator = RecommendationValidator(str(mock_kb_path))
        
        plan_params = {
            'voltage_V': 300.0,
            'current_density_A_dm2': 10.0,
            'frequency_Hz': 500.0
        }
        
        citation_params = {
            'voltage_V': 320.0,
            'current_density_A_dm2': 12.0,
            'frequency_Hz': 500.0
        }
        
        deltas = validator._calculate_param_delta(plan_params, citation_params)
        
        # 300 vs 320: (300-320)/320 = -6.25%
        assert abs(deltas['voltage_V'] - (-6.2)) < 0.1
        
        # 10 vs 12: (10-12)/12 = -16.67%
        assert abs(deltas['current_density_A_dm2'] - (-16.7)) < 0.1
        
        # 500 vs 500: 0%
        assert deltas['frequency_Hz'] == 0.0
    
    def test_determine_match_by_similarity(self, mock_kb_path):
        """测试基于相似度的匹配判断"""
        validator = RecommendationValidator(str(mock_kb_path), similarity_threshold=0.7)
        
        plan = {'system': 'silicate'}
        citations = []
        
        # 高相似度应该匹配
        assert validator._determine_match(0.85, plan, citations) == True
        
        # 低相似度应该不匹配
        assert validator._determine_match(0.5, plan, citations) == False
    
    def test_determine_match_by_content(self, mock_kb_path):
        """测试基于内容的匹配判断"""
        validator = RecommendationValidator(str(mock_kb_path), similarity_threshold=0.9)  # 高阈值
        
        plan = {'system': 'silicate'}
        citations = [
            {
                'text': 'Na2SiO3 electrolyte processed at 300V with 10A/dm2 current density and 500Hz frequency'
            }
        ]
        
        # 即使相似度不高，但内容匹配应该返回True
        assert validator._determine_match(0.5, plan, citations) == True
    
    @patch('scripts.validate_recommendations.kb_search')
    def test_validate_plan_success(self, mock_kb_search, mock_kb_path, mock_search_results):
        """测试方案验证成功"""
        mock_kb_search.return_value = mock_search_results
        
        validator = RecommendationValidator(str(mock_kb_path))
        
        plan = {
            'plan_id': 'test_001',
            'system': 'silicate',
            'voltage': 300,
            'current_density': 10.0
        }
        
        result = validator.validate_plan(plan, topk=3)
        
        assert isinstance(result, ValidationResult)
        assert result.plan_id == 'test_001'
        assert result.system == 'silicate'
        assert result.match_found == True  # 高相似度0.85应该匹配
        assert result.similarity_score == 0.85
        assert len(result.nearest_citations) == 3
        assert result.nearest_citations[0]['source'] == 'test_paper_1.pdf'
        assert isinstance(result.delta_params, dict)
    
    @patch('scripts.validate_recommendations.kb_search')
    def test_validate_plan_failure(self, mock_kb_search, mock_kb_path):
        """测试方案验证失败处理"""
        mock_kb_search.side_effect = Exception("Search failed")
        
        validator = RecommendationValidator(str(mock_kb_path))
        
        plan = {
            'plan_id': 'test_001',
            'system': 'silicate'
        }
        
        result = validator.validate_plan(plan)
        
        assert result.plan_id == 'test_001'
        assert result.match_found == False
        assert result.similarity_score == 0.0
        assert len(result.nearest_citations) == 0
    
    @patch('scripts.validate_recommendations.kb_search')
    def test_validate_batch(self, mock_kb_search, mock_kb_path, sample_plans_csv, mock_search_results):
        """测试批量验证"""
        mock_kb_search.return_value = mock_search_results
        
        validator = RecommendationValidator(str(mock_kb_path))
        
        results, summary = validator.validate_batch(str(sample_plans_csv), topk=3)
        
        # 验证结果
        assert len(results) == 2
        assert all(isinstance(r, ValidationResult) for r in results)
        assert results[0].plan_id == 'test_plan_001'
        assert results[1].plan_id == 'test_plan_002'
        
        # 验证摘要
        assert isinstance(summary, ValidationSummary)
        assert summary.total_plans == 2
        assert summary.matched_plans == 2  # 高相似度都应该匹配
        assert summary.match_rate == 1.0
        assert summary.avg_similarity == 0.85
    
    def test_generate_summary(self, mock_kb_path):
        """测试摘要生成"""
        validator = RecommendationValidator(str(mock_kb_path))
        
        results = [
            ValidationResult(
                plan_id='plan_1', system='silicate', match_found=True,
                similarity_score=0.8, nearest_citations=[{'source': 'paper1.pdf'}],
                delta_params={'voltage_V': 5.0, 'current_density_A_dm2': -10.0},
                query_used='test', validation_time='2024-01-01'
            ),
            ValidationResult(
                plan_id='plan_2', system='zirconate', match_found=False,
                similarity_score=0.4, nearest_citations=[{'source': 'paper2.pdf'}],
                delta_params={'voltage_V': 15.0, 'current_density_A_dm2': 20.0},
                query_used='test', validation_time='2024-01-01'
            )
        ]
        
        summary = validator._generate_summary(results)
        
        assert summary.total_plans == 2
        assert summary.matched_plans == 1
        assert summary.unmatched_plans == 1
        assert summary.match_rate == 0.5
        assert abs(summary.avg_similarity - 0.6) < 0.001  # (0.8 + 0.4) / 2
        assert summary.avg_delta_voltage == 10.0  # (5.0 + 15.0) / 2
        assert summary.avg_delta_current == 15.0  # (10.0 + 20.0) / 2
    
    @patch('scripts.validate_recommendations.kb_search')
    def test_export_results_excel(self, mock_kb_search, mock_kb_path, sample_plans_csv, mock_search_results, temp_dir):
        """测试Excel导出功能"""
        mock_kb_search.return_value = mock_search_results
        
        validator = RecommendationValidator(str(mock_kb_path))
        results, summary = validator.validate_batch(str(sample_plans_csv), topk=3)
        
        output_path = temp_dir / "test_validation_report.xlsx"
        exported_path = validator.export_results(results, summary, str(output_path))
        
        # 验证文件存在
        assert exported_path.exists()
        
        # 如果有openpyxl，验证Excel结构
        try:
            import openpyxl
            wb = openpyxl.load_workbook(exported_path)
            
            # 验证工作表存在
            assert "Summary" in wb.sheetnames
            assert "Matched" in wb.sheetnames or "Unmatched" in wb.sheetnames
            
            # 验证摘要数据
            summary_ws = wb["Summary"]
            assert summary_ws.cell(2, 1).value == "总方案数"
            assert summary_ws.cell(2, 2).value == 2
            
        except ImportError:
            # 如果没有openpyxl，应该生成CSV
            csv_path = output_path.with_suffix('.csv')
            assert csv_path.exists()
    
    def test_export_results_csv_fallback(self, mock_kb_path, temp_dir):
        """测试CSV导出兜底功能"""
        validator = RecommendationValidator(str(mock_kb_path))
        
        # 创建测试数据
        results = [
            ValidationResult(
                plan_id='plan_1', system='silicate', match_found=True,
                similarity_score=0.8, nearest_citations=[
                    {'source': 'paper1.pdf', 'page': 1, 'score': 0.8, 'text': 'test text'}
                ],
                delta_params={'voltage_V': 5.0},
                query_used='test query', validation_time='2024-01-01'
            )
        ]
        
        summary = ValidationSummary(
            total_plans=1, matched_plans=1, unmatched_plans=0,
            match_rate=1.0, avg_similarity=0.8, avg_delta_voltage=5.0,
            avg_delta_current=0.0, most_cited_sources=[('paper1.pdf', 1)],
            validation_time='2024-01-01'
        )
        
        # 模拟openpyxl不可用
        with patch('openpyxl.Workbook', side_effect=ImportError):
            output_path = temp_dir / "test_report.xlsx"
            exported_path = validator.export_results(results, summary, str(output_path))
            
            # 应该生成CSV文件
            assert exported_path.suffix == '.csv'
            assert exported_path.exists()
            
            # 验证CSV内容
            df = pd.read_csv(exported_path)
            assert len(df) == 1
            assert 'match_found' in df.columns
            assert df.iloc[0]['match_found'] == True

class TestIntegration:
    """集成测试"""
    
    @pytest.fixture
    def temp_workspace(self):
        """创建临时工作空间"""
        temp_dir = tempfile.mkdtemp()
        workspace = pathlib.Path(temp_dir)
        
        # 创建目录结构
        (workspace / "datasets" / "index_store").mkdir(parents=True)
        (workspace / "tasks").mkdir(parents=True)
        
        yield workspace
        shutil.rmtree(temp_dir)
    
    def test_full_workflow(self, temp_workspace):
        """测试完整工作流程"""
        # 创建测试用的方案CSV
        plans_data = [
            {
                'plan_id': 'batch_test_001',
                'system': 'silicate',
                'voltage_V': '300',
                'current_density_A_dm2': '10.0',
                'alpha': '0.20',
                'epsilon': '0.80'
            }
        ]
        
        plans_csv = temp_workspace / "tasks" / "test_plans.csv"
        with open(plans_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=plans_data[0].keys())
            writer.writeheader()
            writer.writerows(plans_data)
        
        # 模拟知识库搜索
        with patch('scripts.validate_recommendations.kb_search') as mock_search:
            mock_search.return_value = {
                'results': [
                    {
                        'source': 'fixture_paper.pdf',
                        'page': 1,
                        'score': 0.75,
                        'text': 'Silicate electrolyte MAO process at 300V, 10A/dm2'
                    }
                ]
            }
            
            # 创建验证器并执行验证
            validator = RecommendationValidator(
                str(temp_workspace / "datasets" / "index_store"),
                similarity_threshold=0.6
            )
            
            results, summary = validator.validate_batch(str(plans_csv), topk=3)
            
            # 验证结果
            assert len(results) == 1
            assert results[0].match_found == True
            from dataclasses import asdict
            assert 'match_found' in asdict(results[0])
            
            # 导出报告
            report_path = temp_workspace / "tasks" / "validation_report.xlsx"
            exported = validator.export_results(results, summary, str(report_path))
            
            # 验证报告文件存在
            assert exported.exists()

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
